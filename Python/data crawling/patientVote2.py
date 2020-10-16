import json
import requests
import csv
import time
from multiprocessing import Pool
from openpyxl import load_workbook

header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36"
}
url = "https://m.haodf.com/ndoctor/ajaxshowlist?is2YearsAgo=0&diseasename=%E5%85%A8%E9%83%A8&diseasekey=all"

def main(col):
    wb = load_workbook("E:/python scripts/haodf/doctor4id.xlsx")
    ws = wb.active
    e = 0
    savePath = 'E:/python scripts/haodf/patvote2_' + str(col) + '.csv'
    for row in range(1,9466):
        # try:
            if ws.cell(row, col).value == None:
                break
            doctorid = ws.cell(row, col).value
            print(doctorid)
            params = {
                'doctorId': doctorid,
                'num': 1,
                'size': 20
            }
            try:
                response = requests.get(url, params=params, headers=header)
                jsons = json.loads(response.text)
            except Exception as err:
                print(err)
                time.sleep(2)
                try:
                    response = requests.get(url, params=params, headers=header)
                    jsons = json.loads(response.text)
                except Exception as err:
                    print(err)
                    time.sleep(2)
                    response = requests.get(url, params=params, headers=header)
                    jsons = json.loads(response.text)
            totalPage = jsons.get('data').get('pageInfo').get('totalPage')
            print(totalPage,jsons)
            if totalPage == 0:
                info = (doctorid,)
                print('--正在存储', info)
                with open(savePath, 'a+', newline='', encoding='utf_8_sig') as f:
                    csv_write = csv.writer(f)
                    csv_write.writerow(info)
                    f.close()
            for i in range(1,totalPage+1):
                params = {
                    'doctorId': doctorid,
                    'num': i,
                    'size': 20
                }
                try:
                    response = requests.get(url, params=params, headers=header)
                    print(response, response.text)
                    jsons = json.loads(response.text)
                except Exception as err:
                    print(err)
                    time.sleep(2)
                    try:
                        response = requests.get(url, params=params, headers=header)
                        print(response, response.text)
                        jsons = json.loads(response.text)
                    except Exception as err:
                        print(err)
                        time.sleep(2)
                        response = requests.get(url, params=params, headers=header)
                        print(response, response.text)
                        jsons = json.loads(response.text)
                lists = jsons.get('data').get('list')
                print(lists)
                for list in lists:
                    print(type(list),list)
                    postList = []
                    if list['postCnt'] > '0':
                        postList = list.pop('postList')
                    print(postList)
                    info = (doctorid,) + tuple(list.values())
                    if postList:
                        for post in postList:
                            info += tuple(post.values())
                    print('--正在存储', info)
                    with open(savePath, 'a+', newline='', encoding='utf_8_sig') as f:
                        csv_write = csv.writer(f)
                        csv_write.writerow(info)
                        f.close()
        # except Exception as err:
        #     e += 1
        #     print(e,err)
        #     continue

if __name__=='__main__':
    for col in range(2,5):
        main(col)
    # pool = Pool(3)
    # pool.map(main,range(2,5))
    # main(4)